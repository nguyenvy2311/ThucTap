import time
import numpy as np
import cv2
import os
from PIL import Image, ImageTk

import requests
import base64
import tkinter as tk
from tkinter import messagebox
from threading import *
from openpyxl import *
from datetime import datetime

cap = cv2.VideoCapture(0)
id_off = '5'
id_cam = '6'
token = 'ab6fb41dc1ce579949a93237a0345869'
url = 'https://ai.giaiphapmobifone.vn/api/face'
login_url = 'https://ai.giaiphapmobifone.vn/api/login'
############################## MAIN WINDOW ##############################
class Win1:
    def __init__(self, master, url, id_off, id_cam, token):
        self.id_off = id_off
        self.id_cam = id_cam
        self.token = token
        #print(self.id_off)
        #print(self.id_cam)
        #print(self.token)
        self.master = master
        self.master.title('MobiFone | Hệ thống nhận diện khuôn mặt')
        self.master.resizable(width=False, height=False)
        self.master.geometry("1300x820")
        self.frame = tk.Frame(self.master)
        self.goButton = tk.Button(self.frame, text = "Cấu hình",  command = self.new_window).pack()
        self.lmain = tk.Label(self.frame, compound=tk.CENTER, anchor=tk.CENTER, relief=tk.RAISED)
        self.button = tk.Button(self.frame, text="Quit", command=self._quit)
        self.disp_tf = tk.Label(self.frame, width=38, font=('Arial', 14))
        self.lmain.pack()
        self.button.pack(side=tk.BOTTOM)
        self.disp_tf.pack(pady=5)
        self.frame.pack()
        
        self.show_frame()
        self.tapi()

    def detect_face(self, img):
        self.img = img
        self.face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
        self.faces = self.face_cascade.detectMultiScale(img, 1.2, 5)
       
        if self.faces == ():
            return False
        else:
            for (x, y, w, h) in self.faces:
                self.img = cv2.rectangle(self.img, (x, y), (x + w, y + h), (255, 0, 0), 2)
        
    def show_frame(self):      
        self.ret, self.framee = cap.read()
        self.rgb = cv2.cvtColor(self.framee, cv2.COLOR_BGR2RGB)
        
        self.detect_face(self.rgb)

        self.prevImg = Image.fromarray(self.rgb)
        self.imgtk = ImageTk.PhotoImage(image=self.prevImg)
        self.lmain.imgtk = self.imgtk
        self.lmain.configure(image=self.imgtk)
        self.lmain.after(10, self.show_frame)

    def dis_clear(self):
        self.disp_tf.config(text = '')

    def api(self):
        #Convert to base64  
        self.ret, self.framee = cap.read()
        self.ret, self.buffer = cv2.imencode('.jpg', self.framee)
        self.jpg_as_text = base64.b64encode(self.buffer)
         #API call
        self.mystr = {
            'office_id' : self.id_off,
            'camera_id' : self.id_cam,
            'image' : self.jpg_as_text,
            'temp' : '',
            'token' : self.token
            }
        self.r = requests.post(url, data = self.mystr)

        if self.r.status_code == 200:
            self.res = self.r.json()
            if self.res['data']:
                now = datetime.now() # current date and time
                date_time = now.strftime("%d/%m/%Y, %H:%M:%S")
                self.disp_tf.config(text = 'Tên: ' + self.res['data']['name'] +'\n' + 'Thời gian: '+ date_time)
                self.disp_tf.after(5000, self.dis_clear)
            else:
                self.disp_tf.config(text = 'Khuôn mặt chưa khai báo lên hệ thống !')
                self.disp_tf.after(5000, self.dis_clear)
        else:
            self.disp_tf.config(text = 'Chưa phát hiện khuôn mặt !')
            self.disp_tf.after(5000, self.dis_clear)

    def _quit(self):
        self.master.quit()     # stops mainloop
        self.master.destroy()  # this is necessary on Windows to prevent
                        # Fatal Python Error: PyEval_RestoreThread: NULL tstate

    #Avoid threading error
    def tapi(self):
        self.t = Timer(5.0, self.tapi)
        try:
            self.api()
        except RuntimeError:
            pass
        else:
            self.t.start()

    def new_window(self):
        self.new_window = tk.Toplevel(self.master)
        self.app = Win2(self.new_window)
        
        
############################## LOGIN WINDOW ##############################
class Win2():
    def __init__(self, master):
        self.username = tk.StringVar(value='b1706556')
        self.password = tk.StringVar(value='123456')
        
        self.master = master
        self.master.geometry("550x120")
        self.master.title('Đăng nhập')
        self.master.resizable(False, False)
        self.frame = tk.Frame(self.master)
        
        self.btn = tk.Button(self.frame, text = 'Đăng nhập', command = self.check)
        self.btn.grid(row=5,column=1, ipadx="15", sticky=tk.W)
        self.close_btn = tk.Button(self.frame,text = 'Đóng', command = self.close_window)
        self.close_btn.grid(row=6,column=1, ipadx="30", sticky=tk.W)
        self.frame.pack()

        # create a Form label

        self.usr_label = tk.Label(self.frame, text="Tên đăng nhập")
        self.pas_label = tk.Label(self.frame, text="Mật khẩu")
        
        self.usr_label.grid(row=1, column=0, sticky=tk.W)
        self.pas_label.grid(row=2, column=0, sticky=tk.W)

        self.usr_entry = tk.Entry(self.frame, textvariable = self.username)
        self.pas_entry = tk.Entry(self.frame, textvariable = self.password, show="*")

        self.usr_entry.grid(row=1, column=1, ipadx="100", sticky=tk.W)
        self.pas_entry.grid(row=2, column=1, ipadx="100", sticky=tk.W)        
        
    def check(self):
        self.usr = self.username.get()
        self.pas = self.password.get()
        dta = {
            'username' : self.usr,
            'password' : self.pas,
            }

        self.res = requests.post(login_url, data = dta)
        
        if ((self.usr == '') or (self.pas == '')):
            self.blank()
        else:
            if self.res.status_code == 200:
                self.resp = self.res.json()
                if (self.resp['status'] == 1):
                    token = self.resp['data']['token']
                    
                    self.master.withdraw()
                    self.new_window = tk.Toplevel(self.master)
                    self.app = Win3(self.new_window, self.usr, token) 
                else:
                    self.password_not_recognised()
            else:
                print("user_not_found()")

    def blank(self):
        messagebox.showerror("Thông báo", "Tên đăng nhập và mật khẩu không được để trống!")

    def password_not_recognised(self):
        messagebox.showerror("Thông báo", "Sai tên đăng nhập hoặc mật khẩu!")
            
    def user_not_found():
        messagebox.showerror("Thông báo", "Lỗi cú pháp!")
    
    def close_window(self):
        self.master.destroy()


############################## CONFIG WINDOW ##############################
class Win3():
    def __init__(self, master, usr, token):
        self.usr = usr
        self.token = token
        self.url_server = tk.StringVar(value='https://ai.giaiphapmobifone.vn/api/face')
        self.id_office = tk.StringVar(value='5')
        self.id_camera = tk.StringVar(value='6')
            
        # opening the existing excel file
        self.wb = load_workbook('/Users/nguyenvy/TT/excel.xlsx')

        # create the sheet object
        self.sheet = self.wb.active
        
        self.master = master
        self.master.geometry("550x150")
        self.master.title('Cấu hình')
        self.master.resizable(False, False)
    
        self.frame = tk.Frame(self.master)
        self.logs_btn = tk.Button(self.frame,text = 'Lịch sử', command = self.new_window)
        self.logs_btn.grid(row=0,column=1, sticky=tk.W)
        self.sub_btn = tk.Button(self.frame,text = 'Xác nhận', command = self.submit)
        self.sub_btn.grid(row=9,column=1, sticky=tk.W)
        self.frame.pack()
        
        self.excel()

        # create a Form label

        self.url_label = tk.Label(self.frame, text="url_server")
        self.oid_label = tk.Label(self.frame, text="id_office")
        self.cid_label = tk.Label(self.frame, text="id_camera")
        
        self.url_label.grid(row=1, column=0, sticky=tk.W)
        self.oid_label.grid(row=2, column=0, sticky=tk.W)
        self.cid_label.grid(row=3, column=0, sticky=tk.W)

        self.url_entry = tk.Entry(self.frame, textvariable = self.url_server)
        self.oid_entry = tk.Entry(self.frame, textvariable = self.id_office)
        self.cid_entry = tk.Entry(self.frame, textvariable = self.id_camera)

        self.url_entry.grid(row=1, column=1, ipadx="130", sticky=tk.W)
        self.oid_entry.grid(row=2, column=1, ipadx="130", sticky=tk.W)
        self.cid_entry.grid(row=3, column=1, ipadx="130", sticky=tk.W)

        self.excel()

    def excel(self):
        self.sheet.column_dimensions['A'].width = 50
        self.sheet.column_dimensions['B'].width = 10
        self.sheet.column_dimensions['C'].width = 10

        self.sheet.cell(row=1, column=1).value = "url_server"
        self.sheet.cell(row=1, column=2).value = "id_office"
        self.sheet.cell(row=1, column=3).value = "id_camera"       
        
    def submit(self):
        self.url = self.url_server.get()
        self.off_id = self.id_office.get()
        self.cam_id = self.id_camera.get() 

        self.current_row = self.sheet.max_row
        self.current_column = self.sheet.max_column

        self.sheet.cell(row=self.current_row + 1, column=1).value = self.url_server.get()
        self.sheet.cell(row=self.current_row + 1, column=2).value = self.id_office.get()
        self.sheet.cell(row=self.current_row + 1, column=3).value = self.id_camera.get()

        # save the file
        self.wb.save('/Users/nguyenvy/TT/excel.xlsx')

        url = self.url
        id_off = self.off_id
        id_cam = self.cam_id
        token = self.token
        
        self.master.withdraw()
        
        self.new_window = tk.Toplevel(self.master)
        self.app = Win1(self.new_window, url, id_off, id_cam, token) 
        
        self.master.destroy()

    def new_window(self):
        self.new_window = tk.Toplevel(self.master)
        self.app = Win4(self.new_window, self.usr, token) 
        
        
############################## LOGS WINDOW ##############################
class Win4():
    def __init__(self, master, usr, token):
        self.usr = usr
        self.token = token
        self.master = master
        self.master.geometry("550x260")
        self.master.title('Lịch sử')
        self.master.resizable(False, False)
        self.frame = tk.Frame(self.master)

        self.close_btn = tk.Button(self.frame,text = 'Đóng', command = self.close_window, anchor=tk.CENTER)
        self.close_btn.grid(row=11,column=1, ipadx="20", sticky=tk.W)
        self.frame.pack()

        self.id_label = tk.Label(self.frame, text="Mã checkin ", font='None 12 bold')
        self.name_label = tk.Label(self.frame, text="Họ tên", font='None 12 bold')
        self.time_label = tk.Label(self.frame, text="Thời gian", font='None 12 bold')

        self.id_label.grid(row=0, column=0, ipadx="20", sticky=tk.W)
        self.name_label.grid(row=0, column=1, ipadx="30", sticky=tk.W)
        self.time_label.grid(row=0, column=2, ipadx="50", sticky=tk.W)

        self.id_entry = tk.Label(self.frame)
        self.name_entry = tk.Label(self.frame)          
        self.time_entry = tk.Label(self.frame)
                  
        self.id_entry.grid(row=1, column=0, ipadx="20", sticky=tk.W)
        self.name_entry.grid(row=1, column=1, ipadx="30", sticky=tk.W)
        self.time_entry.grid(row=1, column=2, ipadx="50", sticky=tk.W)  
        url = 'https://ai.giaiphapmobifone.vn/api/logs'
        dta = {
                'username' : self.usr,
                'token' : self.token
                }
        res = requests.post(url, data = dta)
        r = res.json()

        data = r['data']
        ids = []
        names = []
        times = []

        for ar in data:
            ids.append(ar["id_checkin_logs"])
            names.append(ar['first_name'] + ' ' + ar['last_name'])
            times.append(ar['created_at'])
            
        for i in reversed(range(len(data) - 10, len(data))):
            cr_date = times[i]
            cr_date = datetime.strptime(cr_date, '%Y-%m-%d %H:%M:%S')
            cr_date = cr_date.strftime("%d/%m/%Y, %H:%M:%S")
            self.id_entry['text'] =  self.id_entry['text'] + '\n' + str(ids[i]) 
            self.name_entry['text'] =  self.name_entry['text'] + '\n' + names[i]
            self.time_entry['text'] =  self.time_entry['text'] + '\n' + cr_date
          
    def close_window(self):
        self.master.destroy()
############################################################################

############################################################################
def main():
    root = tk.Tk()
    app = Win1(root, url, id_off, id_cam, token)
    root.mainloop()

if __name__ == '__main__':
    main()
